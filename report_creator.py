import openpyxl
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.marker import DataPoint
import json
from common import *
import os.path

parser.add_argument('--clear', nargs='*', default=[], help='sheet preffixes to be cleared')
parser.add_argument('prefix', nargs='?', help='optional prefix for the created sheets')
parser.add_argument('subtract', nargs='?', help='subtract the number from the sheet with this prefix')
args = parse_args()
col_letter = openpyxl.utils.cell.get_column_letter

load_agent_template = lambda: openpyxl.load_workbook(filename = 'template.xlsx')
wb = openpyxl.load_workbook(filename = 'reports/General Dispositions.xlsx')
column = 1
column_names = []
for cell in wb.active['1']:
  column_names.append(cell.value.replace(' ', '').lower())
disposition_names = column_names[1:]
def name_to_column_letter(name):
  return col_letter(col_index(name))
def col_index(name):
  return column_names.index(name)+1
sheet_title = today.strftime('%b %d')
if args.subtract:
  subtract_title = args.subtract + ' ' + sheet_title
  subtract_stat_dict = {}
  subtract_ws = wb[subtract_title]
  for row in subtract_ws.iter_rows(min_row=3):
    if row[0].value:
      subtract_stat_dict[row[0].value] = [int(cell.value) for cell in row[1:]]
else:
  subtract_stat_dict = None

if args.clear:
  cleared_titles = [ prefix + ' ' + sheet_title for prefix in args.clear ]
  print('Cleared sheets:', cleared_titles)
else:
  cleared_titles = []

if args.prefix:
  sheet_title = args.prefix + ' ' + sheet_title

def get_sheet(wb, title, cleared_titles):
  for cleared in cleared_titles:
    if cleared in wb.sheetnames:
      print('Deleting existing sheet: ', cleared)
      del wb[cleared]
  if title in wb.sheetnames:
    return wb[title]
  current = wb.copy_worksheet(wb[wb.sheetnames[0]])
  current.title = title
  return current
def create_sheet(wb, title, cleared_titles):
  for cleared in cleared_titles:
    if cleared in wb.sheetnames:
      print('Deleting existing sheet: ', cleared)
      del wb[cleared]
  if title in wb.sheetnames:
    print('Deleting existing sheet: ', title)
    del wb[title]
  current = wb.copy_worksheet(wb[wb.sheetnames[0]])
  current.title = title
  return current

current = create_sheet(wb, sheet_title, cleared_titles)

def create_chart(ws, title):
  pie = PieChart()
  labels = Reference(ws, min_col=3, max_col=len(column_names), min_row=1)
  data = Reference(ws, min_col=3, max_col=len(column_names), min_row=2)
  pie.add_data(data, from_rows=2, titles_from_data=False)
  pie.set_categories(labels)
  pie.title = title
  pie.width = 20
  pie.height = 18

  series = pie.series[0]

  assigned_labels = []
  def set_label_color(label, color):
    pt = DataPoint(idx=column_names.index(label) - 2)
    pt.graphicalProperties.solidFill = color
    series.dPt.append(pt)
    assigned_labels.append(label)

  set_label_color('voicemail', '009900')
  set_label_color('callback', '00ff88')
  set_label_color('tossback', '00ee00')
  set_label_color('notinterested', '770000')
  set_label_color('renter', 'cc00cc')
  set_label_color('wrongnumber', '770077')
  set_label_color('spanish', 'eeaa00')
  set_label_color('solarhangup', 'aa0000')
  set_label_color('customerhangup', 'ff0000')
  set_label_color('ambassador', '00ffff')
  set_label_color('appointmentwithbill', '00aaff')
  set_label_color('appointmentwithoutbill', '0077ff')
  set_label_color('alreadysolar', '0099bb')

  unassigned_labels = [item for item in column_names if item not in assigned_labels]
  step = int(180 / len(unassigned_labels))
  for i, label in enumerate(unassigned_labels):
    set_label_color(label, hex(i*step + 75)[2:]*3)

  ws.add_chart(pie, 'C15')

row = 3

last_hours = 3
dispositions = request_json('Reports/GetUserDispositionStatsByCampaign', 'POST', data='{"startDate":"'+day_start+'","endDate":"'+day_end+'","campaignID":0,"customerID":"1","userID":"2"}')
for agent in dispositions:
  name = agent['name']
  stats = [None for _ in range(len(disposition_names))]
  for disposition in agent['dispositionStatsList']:
    stats[disposition_names.index(disposition['dispositionName'].replace(' ', '').lower())] = disposition['count']
  if subtract_stat_dict and name in subtract_stat_dict:
    subtract_stats = subtract_stat_dict[name]
    skipped = True
    for i in range(len(disposition_names)):
      stats[i] -= subtract_stats[i]
      if stats[i] > 0:
        skipped = False
    if skipped:
      continue
  path = f'reports/{name}.xlsx'
  if os.path.isfile(path):
    agent_wb = openpyxl.load_workbook(filename = path)
    print(name)
    agent_ws = get_sheet(agent_wb, sheet_title, cleared_titles)
  else:
    agent_wb = load_agent_template()
    agent_ws = agent_wb.active
    agent_ws.title = sheet_title
  current[name_to_column_letter('name')+str(row)] = name
  agent_ws[name_to_column_letter('name')+'2'] = name
  for i, num in enumerate(stats):
    assert num != None, 'None in column ' + column_names[i]
    column = col_letter(i + 2)
    current[column + str(row)] = num
    agent_ws[column + '2'] = num
  create_chart(agent_ws, 'Agent Dispositions')
  agent_wb.save(filename=f'reports/{name}.xlsx')
  row +=1

while True:
  if not current['A'+str(row)].value: break
  for column in range(len(column_names)):
    current[col_letter(column+1)+str(row)] = ''
  row += 1

create_chart(current, 'General Dispositions')

wb.save(filename=f'reports/General Dispositions.xlsx')
